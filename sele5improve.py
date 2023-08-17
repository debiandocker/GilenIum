# coding=utf-8
from selenium import webdriver
import re
from xlutils.copy import copy
import openpyxl
from bs4 import BeautifulSoup
import time
import datetime
import ftplib


# python 3.6 only
# Used to download chromedriver.exe
#from webdriver_manager.chrome import ChromeDriverManager
#driver = webdriver.Chrome(ChromeDriverManager().install())


now_date = datetime.date.today()
now_time = datetime.datetime.strptime(str(now_date), '%Y-%m-%d')
# Please change to your local directory
PATH = r"C:\Users\zebaarap\Documents\TransferToZeba\TransferToZeba\JSA"

class MyException(Exception):
    def __init__(self, *args):
        self.args = args


def LoadComplete(driver):
    STR_READY_STATE = ''
    time_start = time.time()
    while STR_READY_STATE != 'complete':
        time.sleep(0.001)
        STR_READY_STATE = driver.execute_script('return document.readyState')
        time_end = time.time()
        if int(time_end - time_start) > 100:
            raise MyException("Open below page above 100s:")
            return False
    return True


def setup_webdriver_instance():
    # option = webdriver.Chrome()
    # option.add_argument("headless")
    driver = webdriver.Chrome()
    # driver = webdriver.PhantomJS()
    driver.implicitly_wait(300)  # 300
    driver.maximize_window()
    driver.set_page_load_timeout(300)  # 300

    try:
        first_url = 'https://kb.juniper.net/InfoCenter/index?page=content&channel=SECURITY_ADVISORIES&act=login'
        driver.get(first_url)
        print(first_url)
        if not LoadComplete(driver):
            print(first_url)
        else:
            return driver
    except Exception as e:
        print(e)


def scan_current_page(br, page):
    """
    :param br:  this is a machanize instance to open tsb amd jsa main page
    :return: all tsb and jsa now , maybe return dict with tsb/jsa update time mapping
    """
    total_jsa_list = []
    stop_scan = False
    print("Go to page %s to collect JSA number List." % page)
    offset = str(page * 15)
    #jsa_site = r"https://kb.juniper.net/InfoCenter/index?page=content&channel=SECURITY_ADVISORIES&cat=SIRT_1&sort=datemodified&dir=descending&max=1000&batch=15&rss=true&itData.offset=%s&draft=Y" % offset
    jsa_site = r"https://kb.juniper.net/InfoCenter/index?page=content&channel=SECURITY_ADVISORIES&sort=datemodified&dir=descending&max=1000&batch=15&rss=true&itData.offset=%s&draft=Y" % offset
    print(jsa_site)
    try:
        br.get(jsa_site)
        if not LoadComplete(br):
            print(jsa_site)
    except Exception as e:
        print(e)

    published_div = driver.find_elements_by_xpath("//table[@slot='table']//tr//td//div[contains(@class, 'blue')]")
    for temp_div in published_div:
        JSA_NO = temp_div.find_element_by_xpath("../../following-sibling::td[1]").text
        # print(JSA_NO)
        time_string = temp_div.find_element_by_xpath("../../following-sibling::td[3]").text
        if ("minute" in time_string) or ("hour" in time_string):
            total_jsa_list.append(JSA_NO)
            continue

        ret1 = re.search(r"(\d+) day.*", time_string, re.I)
        if ret1:
            days = ret1.group(1)
            # print(days)
            # published_date = (now_time + datetime.timedelta(days=-int(days))).strftime('%Y-%m-%d')
            published_time = now_time + datetime.timedelta(days=-int(days))
            if published_time > last_update_time:
                total_jsa_list.append(JSA_NO)
                continue
            else:
                print("Stop update JSA from " + JSA_NO)
                stop_scan = True
                break

        ret2 = re.search(r"(\w+?) (\d+), (\d+)", time_string, re.I)
        if ret2:
            combined_string = ret2.group(3) + ret2.group(1) + ret2.group(2)
            published_time = datetime.datetime.strptime(combined_string, '%Y%b%d')
            if published_time > last_update_time:
                total_jsa_list.append(JSA_NO)
                continue
            else:
                print("Stop update JSA from " + JSA_NO)
                stop_scan = True
                break

    print("Length of JSA list in page %s is %s" % (page, str(len(total_jsa_list))))
    for JSA_NO in total_jsa_list:
        print(JSA_NO)

    return total_jsa_list, stop_scan


def info_scrapy(jsa_list, current_row):
    table = openpyxl.load_workbook("%s\extra_jsa.xlsx" % PATH)
    sheet = table.active
    row = current_row + 1
    # print(jsa_list)
    for i in jsa_list:
        print(i)
        jsa_url = 'https://kb.juniper.net/InfoCenter/index?page=content&id=%s&cat=&actp=LIST&showDraft=false' % i
        driver.get(jsa_url)
        if not LoadComplete(driver):
            print(jsa_url)
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        jsa_title_hyper_link = '=HYPERLINK("https://kb.juniper.net/%s", "%s")' % (i, i)
        sheet.cell(row, 3).value = jsa_title_hyper_link
        # title = soup.title.text
        title = "  Juniper Networks - " + soup.find(name="p", attrs={"class": "article-name"}).text
        print(title)

        sheet['D%s' % row] = title

        category_list = []
        if soup.find('jnpr-kb-accordion'):
            li1 = soup.find('jnpr-kb-accordion').find('ul').find_all('a')
        else:
            li1 = []
        for j in li1:
            category_list.append(j.text.strip())
        #print(category_list)
        category_string = ",".join(category_list)
        sheet['G%s' % row] = category_string

        if soup.find('jnpr-kb-accordion').find_next_sibling('jnpr-kb-accordion'):
            span_list = soup.find('jnpr-kb-accordion').find_next_sibling('jnpr-kb-accordion').find_all('span')
            available_to = u''
            for span in span_list:
                if re.search(r'Available To', span.text):
                    available_to = re.search(r'Available To:(.*)', span.text).group(1)
            sheet['E%s' % row] = available_to

        if soup.find(text="Product Affected:"):
            product_affect = soup.find(
                text="Product Affected:").next.next.next.text
        else:
            product_affect = u""
        # sheet.write(row, 6, product_affect)
        sheet['F%s' % row] = product_affect

        if soup.find(text="Problem:"):
            problem = soup.find(
                text="Problem:").next.next.next.text
        else:
            problem = u""
        # sheet.write(row, 8, alertDescription)
        sheet['H%s' % row] = problem

        if soup.find(text="Workaround:"):
            workaround = soup.find(text="Workaround:").next.next.next.text
        else:
            workaround = u""
        sheet['I%s' % row] = workaround

        if soup.find(text="Solution:"):
            solution = soup.find(text="Solution:").next.next.next.text
        else:
            solution = u""
        # sheet.write(row, 9, solution)
        sheet["J%s" % row] = solution

        if soup.find(text="Severity Level:"):
            severity = soup.find(text="Severity Level:").next.next.next.text
        else:
            severity = u""
        sheet['K%s' % row] = severity

        if soup.find(text="CVSS Score:"):
            cvss = soup.find(text="CVSS Score:").next.next.next.text
        else:
            cvss = u""
        sheet['M%s' % row] = cvss

        if soup.find(text="CVSS Score:"):
            cvss = soup.find(text="CVSS Score:").next.next.next.text
        else:
            cvss = u""
        sheet['M%s' % row] = cvss

        # pattern = re.compile(r'(prs?.?\d{5,7}|PRs?.?\d{5,7}|Prs?.?\d{5,7})')
        pattern = re.compile(r'(tracked.*\d{5,7}\.)')
        tracking_pr_list_tmp = pattern.findall(problem) + pattern.findall(solution)
        tracking_pr_list = []
        if tracking_pr_list_tmp:
            for pr_seg in tracking_pr_list_tmp:
                pr_list = re.findall(r'\d{5,7}', pr_seg)
                tracking_pr_list.extend(pr_list)

        tracking_pr_list_unique = list(set(tracking_pr_list))
        tracking_pr_string = ", ".join(tracking_pr_list_unique)
        sheet['L%s' % row] = tracking_pr_string
        row += 1

    table.save("%s\extra_jsa.xlsx" % PATH)


def init_extra_file():
    print("Start clean Extra file......")
    table = openpyxl.load_workbook("%s\extra_jsa.xlsx" % PATH)
    sheet = table.active
    sheet.delete_rows(1, sheet.max_row)
    table.save("%s\extra_jsa.xlsx" % PATH)
    print("Extra file is cleaned.")


def load_to_server():
    table = openpyxl.load_workbook("%s\extra_jsa.xlsx" % PATH)
    sheet = table.active
    if (sheet.max_row == 1) and (sheet.max_column == 1):
        print("No JSA update this time!")
    else:
        ### Download
        print("Start download from Server...")
        server_file_name = "jsa.xlsx"
        before_file_name = "%s\jsa_before_%s.xlsx" % (PATH, str(now_date))

        ftp = ftplib.FTP("172.27.101.72")
        ftp.login("lab", "lab123")
        ftp.cwd("/home/lab/Django_project/mysite/media/jsa/")
        try:
            with open(before_file_name, 'wb') as f:
                ftp.retrbinary('RETR %s' % server_file_name, f.write)
                f.close()
                print("Download finished.")
        except Exception as e:
            print('FTP Download error:', e)

        ### Add delta data
        # moving
        print("Start add new JSA......")
        after_file_name = "%s\jsa_after_%s.xlsx" % (PATH, str(now_date))
        new_table = openpyxl.load_workbook(before_file_name)
        new_sheet = new_table.active
        new_sheet.insert_rows(2, sheet.max_row)
        add_list = []
        update_list = []
        for i in range(1, sheet.max_row + 1):
            update_flg = False
            JSA_NO = str(sheet.cell(i, 3).value.split(',')[1]).strip().strip(')').strip('"')
            # Add one delta line
            for j in range(1, sheet.max_column + 1):
                new_sheet.cell(i + 1, j).value = sheet.cell(i, j).value

            for k in reversed(range(2, new_sheet.max_row + 1)):
                if (JSA_NO in str(new_sheet.cell(k, 3).value)) and (i + 1 != k):
                    update_flg = True
                    update_list.append(JSA_NO)
                    print("%s already in Line %d, should be removed" % (JSA_NO, k - sheet.max_row))
                    new_sheet.delete_rows(k)
            if not update_flg:
                add_list.append(JSA_NO)

        new_table.save(after_file_name)
        print("Finish add new JSA.")
        print("Update JSA:")
        for i in update_list:
            print(i)
        print("New add JSA:")
        for i in add_list:
            print(i)

        # Upload
        print("Start upload to Server...")
        tar_file_name = "jsa_%s.xlsx" % str(now_date)
        try:
            with open(after_file_name, 'rb') as fp:
                res = ftp.storbinary("STOR " + tar_file_name, fp)
                if not res.startswith('226 Transfer complete'):
                    print('Upload failed')
                else:
                    print("Upload finished.")
                    fp.close()
        except ftplib.all_errors as e:
            print('FTP Upload error:', e)
        ftp.quit()


def get_last_date():
    with open("%s\JSA_last_update_date.txt" % PATH, "r") as f1:
        for line in f1:
            last_update_date = line.strip()
    return last_update_date


def set_last_date():
    # current_time = datetime.datetime.now()
    # current_date_string = current_time.strftime("%Y-%m-%d")
    with open("%s\JSA_last_update_date.txt" % PATH, "w") as f2:
        f2.write("%s" % now_date)
    print("Now JSA update date is changed to : %s " % now_date)


########### Main Function
# Get last update time
last_update_date = get_last_date()
print("JSA update from date:" + last_update_date)
last_update_time = datetime.datetime.strptime(last_update_date, '%Y-%m-%d')

# clean the content of the extra_jsa.xlsx file
init_extra_file()

# login to the first URL
driver = setup_webdriver_instance()
# Please change to your userid
driver.find_element_by_id("idp-discovery-username").send_keys("test@juniper.net")
if not driver.find_element_by_xpath("//input[@id='input6']").is_selected():
    driver.find_element_by_xpath("//input[@id='input6']/following-sibling::label").click()
# preceding-sibling

#     remember_me.click()
driver.find_element_by_id('idp-discovery-submit').click()
if not LoadComplete(driver):
    print("Input first use_name timeout!")
# Please change to your userid
driver.find_element_by_id('i0116').send_keys('test@juniper.net')
driver.find_element_by_id('idSIButton9').click()
if not LoadComplete(driver):
    print("Input second use_name timeout!")
# Please change to your password
driver.find_element_by_id('i0118').send_keys('password')
# make sure to click the submit button
time.sleep(8)
driver.find_element_by_id('idSIButton9').click()
# driver.find_element_by_xpath("//input[@id='idSIButton9']/..").click()
# wait for phone confirm
# driver.find_element_by_xpath('//a[@href="'+url+'"]')
while not driver.find_element_by_xpath('//a[contains(@href,"logout")]'):
    time.sleep(2)

row = 0
stop_scan = False
# First 4 pages is enough since we run it once a week
for page in range(0, 5):
    if not stop_scan:
        jsa_list, stop_scan = scan_current_page(driver, page)
        info_scrapy(jsa_list, row)
        row += len(jsa_list)

# Change the file on server
load_to_server()
# Update the last update date
set_last_date()
